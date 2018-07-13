#!/usr/bin/env python3

"""
This program takes an export Excel file and
tabulates TC counts Open and Closed, by screening and priority,
and by date closed.
File type must be .xlsx, older .xls are not supported by openpyxl.
Date of last change:
Jonathan McDonald
Args:
Returns:
Raises:
"""

# Import and include
import re
import openpyxl
import pprint
import datetime
# import os.path
from os.path import isfile as checkFile
# import excelBeanCounter
# from excelBeanCounter import rowCounter

# Let me know I have started
# open downloaded TSM hull export for today
# ask me the file name
##try:
##    myDay = datetime.date.today()
##    print('myDay: ' + str(myDay))
##    today = str(myDay)
##    print('today s1: ' + today)
##    today = today.replace('-', '')
##    print('today s2: ' + today)
##
##    myHullNum = input('What is the Hull Number? ')
##    curTSM_xlsx = today + '_1 LPD ' + myHullNum + ' ALL bean bu.xls'
##    print('The file name is: ' + curTSM_xlsx)
##    # Open handles and objects
##    wb = openpyxl.load_workbook(curTSM_xlsx)  # must be in same directory or bat location
##except:
##    curTSM_xlsx = input('What is the name of the TSM Excel export file? ')
##    print('The file name is: ' + curTSM_xlsx)
##    # Open handles and objects
##    wb = openpyxl.load_workbook(curTSM_xlsx)  # Excel file must be in same directory or bat location


# Parse out TC to cls,hull,d,s,p
def parseTCnum(tc_num=None):
    # Get the trial card number off of the current page with reg-ex
    # Trial Card:ABC0000DE-FG000101
    # two to four letters (\w){2,4}
    # four numbers (\d){4}
    # two to four letters (\w){2,4} and optional one number (\d)?
    # one hyphen -
    # two letters (\w){2}
    # six numbers (\d){6}
    reTCnum = re.compile(r'(\w){2,4}(\d){4}(\w){2,4}(\d)?-(\w){2}(\d){6}')
    moTCnum = reTCnum.search(tc_num)  # Passed in trial card number
    if moTCnum is None:
        # No Matched Object Trial Card number passed in
        return None  # Exit with no data.
    else:
        fullTCnum = moTCnum.group()  # matched object returned
        curTCdept = fullTCnum[-8:-6]  # slice out the department string
        # Self check that what was sliced is letters not numbers
        reTCdept = re.compile(r'(\w){2}')
        moTCdept = reTCdept.search(curTCdept)
        if moTCdept is None:
            # Incomplete trial card number passed in
            return None  # Exit with no data.
        else:
            curTCdept = tc_num[-8:-6]  # slice out the literal string "DP"
        return curTCdept


# Row count function
def rowCount(row, bean, dt_BT=None, dt_AT=None):
    # Reset selectors booleans
    myPri_star = False
    myPri_1s = False
    myPri_1 = False
    myPri_2s = False
    myPri_2 = False
    myPri_3s = False
    myPri_other = False

    # Each row in the spreadsheet has data for one Trial Card
    dsp = sheet['A' + str(row)].value  # 'Trial Card #'
    dept = parseTCnum(dsp)  # 'Department from the Trial Card #'
    star = sheet['B' + str(row)].value  # 'Star'
    pri = sheet['C' + str(row)].value  # 'Pri'
    safe = sheet['D' + str(row)].value  # 'Saf'
    scrn = sheet['E' + str(row)].value  # 'Scrn'
    ac1 = sheet['F' + str(row)].value  # 'Act 1'
    ac2 = sheet['G' + str(row)].value  # 'Act 2'
    stat = sheet['H' + str(row)].value  # 'Status'
# actkn not currently used
    # actkn = sheet['I' + str(row)].value  # 'Action Taken'
    dt_disc = sheet['J' + str(row)].value  # 'Date Discovered'
    # Reformat dates from Oracle(-) to Microsoft(/)
    if dt_disc is not None:
        # date_disc = datetime.date(dt_disc, '%Y-%m-%d')
        dt_disc = dt_disc.replace('-', '/')
        pass
    dt_close = sheet['K' + str(row)].value  # 'Date Closed'
    # Reformat dates from Oracle(-) to Microsoft(/)
    if dt_close is not None:
        # date_close = datetime.date(dt_close, '%Y-%m-%d')
        dt_close = dt_close.replace('-', '/')
        pass
# trial_ID and event currently not used
    # trial_ID = sheet['L' + str(row)].value  # 'Trial ID'
    # event = sheet['M' + str(row)].value  # 'Event'

    # Combine singleton values
    # May turn this off and not combine screening codes
    if (ac2 != '') and (ac2 is not None):
        scrngs = scrn + '/' + ac1 + '/' + ac2
    else:
        scrngs = scrn + '/' + ac1

    # Check for Starred Cards
    if (star == 'STAR') or (star == 'star') or (star == '*'):
        myPri_star = True
    elif (star == '') or (star is None):
        # No Starred Cards in row
        if int(pri) == 1:
            if safe == 'S':
                myPri_1s = True
            elif safe != 'S':
                myPri_1 = True
            else:
                # Un-captured value in field
                pass
        elif int(pri) == 2:
            if safe == 'S':
                myPri_2s = True
            elif safe != 'S':
                myPri_2 = True
            else:
                # Un-captured value in field
                pass
        elif int(pri) == 3:
            if safe == 'S':
                myPri_3s = True
            elif safe != 'S':
                myPri_other = True
            else:
                # Un-captured value in field
                pass
        else:
            # Un-captured value in field
            myPri_other = True
            print('Row read error Priority in not Valid. Row: ' + str(row) + ' TC Number: ' + dsp)
    else:
        # No Starred Cards in row, catch all
        pass

## Commented out to stop double counting Stared cards
##    if int(pri) == 1:
##        if safe == 'S':
##            myPri_1s = True
##        elif safe != 'S':
##            myPri_1 = True
##        else:
##            # Un-captured value in field
##            pass
##    elif int(pri) == 2:
##        if safe == 'S':
##            myPri_2s = True
##        elif safe != 'S':
##            myPri_2 = True
##        else:
##            # Un-captured value in field
##            pass
##    elif int(pri) == 3:
##        if safe == 'S':
##            myPri_3s = True
##        elif safe != 'S':
##            myPri_other = True
##        else:
##            # Un-captured value in field
##            pass
##    else:
##        # Un-captured value in field
##        myPri_other = True
##        print('Row read error Priority in not Valid. Row: ' + str(row) + ' TC Number: ' + dsp)

    # Make sure the key(s) for these dictionaries exist.
    # The .setdefault() checks if the key exists, if not it creates
    # with default passed value otherwise it will do nothing.
    myBean = str(bean)
    # By bean type, status and priority count
    tc_Status.setdefault(myBean, {})
    tc_Status[myBean].setdefault(stat, {'Pri STARRED': 0, 'Pri 1S': 0, 'Pri 1': 0, 'Pri 2S': 0, 'Pri 2': 0, 'Pri 3S': 0, 'Pri OTHER': 0, 'Total': 0})
    # By bean type, status, screen and screening count
    tc_Stat_Scrn.setdefault(myBean, {})
    tc_Stat_Scrn[myBean].setdefault(stat, {})
    tc_Stat_Scrn[myBean][stat].setdefault(scrn, {})
    tc_Stat_Scrn[myBean][stat][scrn].setdefault(scrngs, {'Pri STARRED': 0, 'Pri 1S': 0, 'Pri 1': 0, 'Pri 2S': 0, 'Pri 2': 0, 'Pri 3S': 0, 'Pri OTHER': 0, 'Total': 0})
    # By bean type, department and priority count
    tc_Depart.setdefault(myBean, {})
    tc_Depart[myBean].setdefault(dept, {'Pri STARRED': 0, 'Pri 1S': 0, 'Pri 1': 0, 'Pri 2S': 0, 'Pri 2': 0, 'Pri 3S': 0, 'Pri OTHER': 0, 'Total': 0})
    # By bean type, date closed and date closed count
    # Check for empty or None date closed
    if (dt_close == '') or (dt_close is None):
        # Don't count, card still open
        # This could be used as a counting validation
        pass
    else:
        # By bean type, date closed and date closed count
        tc_DateClosed.setdefault(myBean, {})
        # Check if Date Closed is already in the dictionary
        tc_DateClosed[myBean].setdefault(dt_close, {})
        tc_DateClosed[myBean][dt_close].setdefault('Closed Count', 0)
        # Increment the date closed count plus 1
        tc_DateClosed[myBean][dt_close]['Closed Count'] += 1

        # Build days from trial
        if dt_BT is not None:
            tc_DateClosed[myBean][dt_close].setdefault('Days from BT', None)
            dateClosed = datetime.datetime.strptime(dt_close, '%Y/%m/%d')
            date_BT = datetime.datetime.strptime(dt_BT, '%Y/%m/%d')
            date_Delta = (dateClosed - date_BT).days
            tc_DateClosed[myBean][dt_close]['Days from BT'] = date_Delta
        else:
            # BT date not given
            pass
        if dt_AT is not None:
            tc_DateClosed[myBean][dt_close].setdefault('Days from AT', None)
            dateClosed = datetime.datetime.strptime(dt_close, '%Y/%m/%d')
            date_AT = datetime.datetime.strptime(dt_AT, '%Y/%m/%d')
            date_Delta = (dateClosed - date_AT).days
            tc_DateClosed[myBean][dt_close]['Days from AT'] = date_Delta
        else:
            # AT date not given
            pass

    # Update the Status and Screening dictionary's counts
    tc_Status[myBean][stat]['Total'] += 1
    tc_Stat_Scrn[myBean][stat][scrn][scrngs]['Total'] += 1
    # Update the Priority counts
    if myPri_star is True:
        tc_Status[myBean][stat]['Pri STARRED'] += 1
        tc_Stat_Scrn[myBean][stat][scrn][scrngs]['Pri STARRED'] += 1
    elif myPri_1s is True:
        tc_Status[myBean][stat]['Pri 1S'] += 1
        tc_Stat_Scrn[myBean][stat][scrn][scrngs]['Pri 1S'] += 1
    elif myPri_1 is True:
        tc_Status[myBean][stat]['Pri 1'] += 1
        tc_Stat_Scrn[myBean][stat][scrn][scrngs]['Pri 1'] += 1
    elif myPri_2s is True:
        tc_Status[myBean][stat]['Pri 2S'] += 1
        tc_Stat_Scrn[myBean][stat][scrn][scrngs]['Pri 2S'] += 1
    elif myPri_2 is True:
        tc_Status[myBean][stat]['Pri 2'] += 1
        tc_Stat_Scrn[myBean][stat][scrn][scrngs]['Pri 2'] += 1
    elif myPri_3s is True:
        tc_Status[myBean][stat]['Pri 3S'] += 1
        tc_Stat_Scrn[myBean][stat][scrn][scrngs]['Pri 3S'] += 1
    elif myPri_other is True:
        tc_Status[myBean][stat]['Pri OTHER'] += 1
        tc_Stat_Scrn[myBean][stat][scrn][scrngs]['Pri OTHER'] += 1
    else:
        print('Row read error with Status of Open, Priority in not Valid. '
                    'Row: ' + str(row) + ' TC Number: ' + dsp)
        pass


# Get and test the name of Excel file
fileNotFound = True
while fileNotFound is True:
    curTSM_xlsx = input('\nWhat is the name of the TSM Excel export file? ')
    # print('The file name is: ' + curTSM_xlsx)
    if checkFile(curTSM_xlsx):
        fileNotFound = False
    else:
        print('File not found, please re-enter the file name.\n'
              'File must be in the Python root directory.\n')

print('Excel file name: ' + curTSM_xlsx)

# Get the current Hull Number
myHullNum = input('\nWhat is the Hull Number?\n'
                  'Example: 17: ')
print('Hull number is: ' + myHullNum)

# Build Events List
runEvents = input('\nRun reports by Event?\n'
                  'Y or N: ')
runEvents = runEvents.upper()
if runEvents == 'Y':
    userEvents = input('What Events are being counted?\n'
                    'Example: AT,BT,FCT: ')
    userEvents = userEvents.upper()
    events = userEvents
    curReportEvents = [item for item in userEvents.split(',') if item.strip()]
    runByEvents = True
else:
    # Presumed no, pass
    runByEvents = False
    pass
print(curReportEvents)

# Build INSURV List
runTID = input('\nRun reports by Trial ID?\n'
               'Y or N: ')
runTID = runTID.upper()
if runTID == 'Y':
    userTID = input('What INSURV Events are being counted?\n'
                    'Example: C,F or C+: ')
    userTID = userTID.upper()
    curReportTrial_ID = [item for item in userTID.split(',') if item.strip()]
    runByTrial_ID = True
else:
    # Presumed no, pass
    runByTrial_ID = False
    pass
print(curReportTrial_ID)
trial_ID = curReportTrial_ID

# Run by INSURV Department
runDept = input('\nRun reports by INSURV Department?\nY or N: ')
runDept = runDept.upper()
if runDept == 'Y':
    runByDept = True
else:
    # Presumed no, pass
    runByDept = False
    pass
print('Run by Department: ' + str(runByDept))

# Get BT Date
runBT = input('\nCount closure from BT Trial?\nY or N: ')
runBT = runBT.upper()
if runBT == 'Y':
    curFromBT_Date = input('What is the date of the BT Trial?\n'
        'Example: yyyy/mm/dd: ')
    print(curFromBT_Date)
    runFromBT_Date = True
else:
    # Presumed no
    curFromBT_Date = None
    runFromBT_Date = False

userBT = curFromBT_Date

# Get AT Date
runAT = input('\nCount closure from AT Trial?\n'
    'Y or N: ')
runAT = runAT.upper()
if runAT == 'Y':
    curFromAT_Date = input('What is the date of the AT Trial?\n'
        'Example: yyyy/mm/dd: ')
    runFromAT_Date = True
    print(curFromAT_Date)
else:
    # Presumed no
    curFromAT_Date = None
    runFromAT_Date = False

userAT = curFromAT_Date

# Open handles and objects
# Excel file must be in same directory or bat location
wb = openpyxl.load_workbook(curTSM_xlsx)

print('\nOpening workbook...')
try:
    # The default is 'TSM EXPORT'
    sheet = wb.get_sheet_by_name('TSM EXPORT')
except:
    mySheetName = input('\nWhat is the name of the sheet? ')
    sheet = wb.get_sheet_by_name(mySheetName)
    print(sheet)

# Initialize the empty dictionary's
tc_Status = {}  # By bean type, status and priority count
tc_Stat_Scrn = {}  # By bean type, status, screen and screening count
tc_DateClosed = {}  # By bean type, date closed and date closed count
tc_Depart = {}  # By bean type, department and priority count

# Let me know I am reading and looping through file
print('Reading rows...')

for row in range(2, sheet.max_row + 1):
    # Check for Header, Last Row or empty values
    if ((sheet['A' + str(row)].value != 'Trial Card No') or
       (sheet['A' + str(row)].value != 'Trial Card #') or
       (sheet['A' + str(row)].value != 'FOR OFFICIAL USE ONLY') or
       (sheet['A' + str(row)].value != '')
       (sheet['A' + str(row)].value is not None)):
        # Check for empty TC Status values
        if ((sheet['H' + str(row)].value != '') or
            (sheet['H' + str(row)].value is not None)):

            # Check if Event is in current report range
            if sheet['M' + str(row)].value is not None:
                # Single value in a list of values
                if sheet['M' + str(row)].value in curReportEvents:
                    # Process row
                    rowCount(row, events, userBT, userAT)
                    # Stop e iteration of curReportEvents, don't multi count single row
                    pass
                else:
                    # Current e iteration in the current report events list is not in the current row
                    pass
            else:
                # cell is empty
                pass

            # Check if Event ID is in current report range
            if sheet['L' + str(row)].value is not None:
                # Multi values in a list of values
                for t in curReportTrial_ID:
                    if t in sheet['L' + str(row)].value:
                        # Process row
                        rowCount(row, trial_ID, userBT, userAT)
                        # Stop t iteration of curReportTrial_ID, don't multi count single row
                        break
                    else:
                        # Current t iteration in the current report trial ID list is not in the current row
                        continue
            else:
                # cell is empty
                pass

    elif ((sheet['A' + str(row)].value == 'Trial Card No') or
          (sheet['A' + str(row)].value == 'Trial Card #')):
        print('Header row found, pass.')
        pass
    elif sheet['A' + str(row)].value == 'FOR OFFICIAL USE ONLY':
        print('Exported fouo found, pass')
        pass
    elif sheet['A' + str(row)].value == '':
        print('Empty cell found, pass')
        pass
    else:
        # Un-captured value in field
        pass


# Open a  file and write the contents of the dictionaries to it.
print('Writing results...')

resultFile = open('LPD ' + myHullNum + ' Bean Data.py', 'w')
# pprint is python print, and formats as valid python code and structure
resultFile.write('tc_Status = ' + pprint.pformat(tc_Status))
resultFile.write('\ntc_Stat_Scrn = ' + pprint.pformat(tc_Stat_Scrn))

if (runFromAT_Date is not False) or (runFromBT_Date is not False):
    resultFile.write('\ntc_DateClosed = ' + pprint.pformat(tc_DateClosed))
else:
    # no dates given, pass
    pass

if runByDept is not False:
    resultFile.write('\ntc_Depart = ' + pprint.pformat(tc_Depart))
else:
    # no dates given, pass
    pass

# Dispose of object
resultFile.close()

# Completed
print('Bean dictionary completed.')

# End Of Line
hold = input('Press any key to exit.')

# End Of Line
print('Good Bye')
